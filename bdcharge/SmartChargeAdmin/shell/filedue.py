# coding:utf-8
# -*- coding: utf-8 -*-
# Create your tests here.

import os
import openpyxl

def all_path(dirname):
    result = []#所有的文件
    for maindir, subdir, file_name_list in os.walk(dirname):
        # print("1:",maindir) #当前主目录
        # print("2:",subdir) #当前主目录下的所有目录
        # print("3:",file_name_list)  #当前主目录下的所有文件
        for filename in file_name_list:
            if '2019' not in filename or '.DZ' not in filename:
                continue
            apath = os.path.join(maindir, filename)#合并成一个完整路径
            result.append(apath)
    return result


def duefile():
    res = all_path("E:\\其他\\irs\\郑汴热力\\ftpzbrl")

    # 创建excel文件
    wb = openpyxl.Workbook()  # 创建文件对象
    ws = wb.active  # 获取第一个sheet, 激活 worksheet
    row = 1
    ws.cell(row=row, column=1).value = '交易日期'
    ws.cell(row=row, column=2).value = '用户编号'
    ws.cell(row=row, column=3).value = '交易金额'

    for item in res:
        print(item)
        with open(item, "r") as f:
            for line in f:
                line=line.strip('\r').strip('\n')
                if len(line) < 3:
                    continue
                # print("读取文件内容：",line)
                userid=line.split('|')[3]
                money=line.split('|')[4]
                date=line.split('|')[5]

                row = row + 1
                print('ddddddd--',row, date,userid,money)
                ws.cell(row=row, column=1).value = date
                ws.cell(row=row, column=2).value = userid
                ws.cell(row=row, column=3).value = money
                # print(item)
    wb.save('E:\\zbrl.xlsx')  # 保存excel文件
    wb.close()
duefile()


def duefile2():


    # 创建excel文件
    wb = openpyxl.Workbook()  # 创建文件对象
    ws = wb.active  # 获取第一个sheet, 激活 worksheet
    row = 1
    ws.cell(row=row, column=1).value = '交易日期'
    ws.cell(row=row, column=2).value = '用户编号'
    ws.cell(row=row, column=3).value = '交易金额'

    with open('E://2019.DZ', "r") as f:
        for line in f:
            line=line.strip('\r').strip('\n')
            if len(line) < 3:
                continue
            # print("读取文件内容：",line)
            userid=line.split('|')[3]
            money=line.split('|')[4]
            date=line.split('|')[5]

            row = row + 1
            print('ddddddd--',row, date,userid,money)
            ws.cell(row=row, column=1).value = date
            ws.cell(row=row, column=2).value = userid
            ws.cell(row=row, column=3).value = money
            # print(item)
    wb.save('E:\\zmrl.xlsx')  # 保存excel文件
    wb.close()
# duefile2()

