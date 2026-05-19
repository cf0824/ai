#根据库房提供的excel,修改基础资料的规格型号，名称，。。


# -*- coding: gb2312 -*-
import openpyxl
import pymssql
file_name="C:\\Users\\Administrator\\Desktop\\原材料规格命名统计.xlsx"
def duemain():

    wb = openpyxl.load_workbook(filename=file_name,read_only=True)
    ws = wb.active

    sqlserver_conn = pymssql.connect(server='192.168.2.18',
                                     user='sa',
                                     password='luyao123KEJI',
                                     database="db_18",
                                     timeout=20,
                                     autocommit=True)  # sqlserver数据库链接句柄
    cursor = sqlserver_conn.cursor()  # 获取光标

    # 获取行列数
    row = ws.max_row
    column = ws.max_column
    print('row=',row)
    print('column=',column)

    start_row = 2; #从第二行开始读取数据
    now_row=0;
    total_row=row; #总行数
    for i in range(start_row,total_row):
        prd_no = ws.cell(row=i, column=2).value
        prd_name = ws.cell(row=i, column=3).value
        prd_spc = ws.cell(row=i, column=4).value
        prd_mrk = ws.cell(row=i, column=5).value
        prd_spc_ak = ws.cell(row=i, column=7).value
        now_row = now_row + 1
        print('当前正在处理', now_row, prd_no, prd_name)
        # print(now_row, prd_no, prd_name, prd_spc, prd_mrk, prd_spc_new)
        # if now_row > 300:
        #     break;

        sql = "SELECT prdt.prd_no,prdt.name,prdt.SNM, prdt.INV_NAME, prdt.spc, prdt.mrk,  " \
              "(select name from mark where mark_no=prdt.mrk) as mark_name FROM  prdt " \
              "WHERE  prdt.prd_no='%s'" % (prd_no)
        # print(sql)
        cursor.execute(sql)
        row_prd=cursor.fetchone()
        if not row_prd:
            print('品号不存在：', now_row, prd_no, prd_name)
            break;
        if (row_prd[6] and not prd_mrk) or (not row_prd[6] and prd_mrk):
            print('品牌不一致1：', now_row, prd_no, prd_mrk, "!=", row_prd[0], row_prd[6])
        elif not row_prd[6] and not prd_mrk:
            pass
        elif row_prd[6]!=prd_mrk:
            print('品牌不一致2：', now_row, prd_no, prd_mrk,"!=", row_prd[0], row_prd[6] )

        if prd_spc_ak:
            prd_spc_new=prd_spc+'('+prd_spc_ak+')'
        else:
            prd_spc_new=prd_spc
        #更新数据库
        sql="update  prdt set name='%s', spc='%s' where prd_no='%s'" % (prd_name, prd_spc_new, prd_no)
        # print(sql)
        try:
            cursor.execute(sql)
        except Exception as ex:
            print(sql)
            print(ex)
            break;
        sqlserver_conn.commit()

    wb.close()
    sqlserver_conn.close()



duemain()
