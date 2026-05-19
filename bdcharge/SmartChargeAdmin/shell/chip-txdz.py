#通过采购给的excel补全焊点数量, 不支持xls格式

import openpyxl
import pymysql
import xlwt

file_name="C:\\Users\\Administrator\\Desktop\\芯片mac对应测试结果.xlsx"
newfile_name="C:\\Users\\Administrator\\Desktop\\芯片mac对应测试结果-result.xlsx"

def duemain():
    try:
        wb = openpyxl.load_workbook(filename=file_name,read_only=False)
        ws = wb.active



        # 连接数据库
        try:
            mysql_conn = pymysql.Connect(
                host='192.168.2.174',
                port=3306,
                user='lqkj',
                passwd='LQkj666_2019',
                db='lqkj_db',
                charset='utf8')
            cur = mysql_conn.cursor()
        except Exception as e:
            print("MySql数据库连接失败!" + str(e))
            return -1

        # 获取行列数
        row = ws.max_row
        column = ws.max_column
        print('row=',row)
        print('column=',column)

        start_row = 2; #从第二行开始读取数据
        total_row=188; #总行数
        ChipID_List=[]
        for i in range(start_row,total_row):
            chipID = ws.cell(row=i, column=5).value
            if chipID:
                chipID = str(chipID).split('0x')[1]
                print(chipID, i)
                if chipID not in ChipID_List:
                    ChipID_List.append(chipID)
                else:
                    style = xlwt.XFStyle()
                    pattern = xlwt.Pattern()
                    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
                    pattern.pattern_fore_colour = 2  # 红色
                    style.pattern = pattern
                    ws.write(i, 8, '重复', style)

                sql="select Test_Result, Batch_Num from yw_project_product_test_info where chip_id='%s'" % (chipID)
                print(sql)
                cur.execute(sql)
                row = cur.fetchone()
                if row:
                    ws.cell(row=i, column=6, value=row[0])
                    ws.cell(row=i, column=7, value=row[1])
                else:
                    ws.cell(row=i, column=6, value='无')
                    ws.cell(row=i, column=7, value='无')
            # if i>5:
            #     break
        wb.save(newfile_name)
        wb.close()
        mysql_conn.close()
    except Exception as e:
        print(e)
        wb.close()
        mysql_conn.close()


duemain()
