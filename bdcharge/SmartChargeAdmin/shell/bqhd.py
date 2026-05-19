#通过采购给的excel补全焊点数量, 不支持xls格式

import openpyxl
import pymssql
file_name="C:\\Users\\Administrator\\Desktop\\添加焊点.xlsx"
def duemain():
    try:
        wb = openpyxl.load_workbook(filename=file_name,read_only=True)
        ws = wb.active

        sqlserver_conn = pymssql.connect(server='192.168.2.18',
                                         user='sa',
                                         password='luyao123KEJI',
                                         database="db_18",
                                         timeout=20,
                                         autocommit=True)  # sqlserver数据库链接句柄
        cursor = sqlserver_conn.cursor()  # 获取光标

        # 获取行列数
        row = ws.max_row
        column = ws.max_column
        print('row=',row)
        print('column=',column)

        start_row = 2; #从第二行开始读取数据
        total_row=139; #总行数
        for i in range(start_row,total_row):
            prd_no = ws.cell(row=i, column=2).value
            upr_tp = ws.cell(row=i, column=4).value
            if upr_tp==None:
                continue
            if upr_tp=='' or int(upr_tp)==0:
                continue
            # print(prd_no, upr_tp)
            sql="update  prdt set UPR_TP='%s' where prd_no='%s'" % (upr_tp, prd_no)
            print(sql)
            cursor.execute(sql)
            sqlserver_conn.commit()

        wb.close()
        sqlserver_conn.close()
    except Exception as e:
        print(e)
        wb.close()
        sqlserver_conn.close()


duemain()
