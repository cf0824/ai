#coding:utf-8
# -*- coding: utf-8 -*-

#######################################################################
#上传excel文件入库操作。异步操作。
#######################################################################
import os
import time
import datetime
import pymysql
import logging
import xlwt
import xlrd
import openpyxl

#日志初始化,日志名称
localhome="/home/admin/lqkj_admin/"
global logger, fh
logger = logging.getLogger(str(os.getpid()))
def loger_init( ):
    #日志初始化
    now_time = datetime.datetime.now().strftime('%Y%m%d')
    logger = logging.getLogger(str(os.getpid()))
    logger.setLevel(logging.INFO)
    logname = os.path.basename(__file__).split('.py')[0]

    if os.path.exists(localhome+'log/'):
        log_file_temp = localhome+'log/'+logname+'_'+now_time+'.log'
    else:
        log_file_temp = logname + '_' + now_time + '.log'
    # print(log_file_temp)
    fh = logging.FileHandler(log_file_temp)  # 定义一个写文件的handler
    fh.setLevel(logging.INFO)  # 设置写文件的等级
    fh_formatter = logging.Formatter(
        '[%(levelname)-5s] [%(filename)-12s line:%(lineno)-4d] [%(asctime)s] [%(process)-7d] [%(message)s]')  # 设置输出格式
    fh.setFormatter(fh_formatter)  # 将输出格式设置给handler
    #print('public',logger)
    if  not logger.handlers:
        logger.addHandler(fh)  # 将handler加入logger
    return logger, fh

#重新连接数据库
def dbconnect(reconnect=True):
    global mysql_conn
    try:
        mysql_conn.ping(reconnect=True)
    except:
        try:
            mysql_conn = pymysql.Connect(
                host='192.168.2.174',
                port=3306,
                user='lqkj',
                passwd='LQkj666_2019',
                db='lqkj_db',
                charset='utf8', autocommit=1)
        except Exception as e:
            print("MySql数据库连接失败!" + str(e))
            mysql_conn.close()
            return False

    return mysql_conn

#导入某某数据时，需要特殊处理
def fileup_proc_191( myfielddict ):
    log, fh = loger_init()
    log.info('----------------------提供给客户芯片ID校验-----------------------')
    log.info( str(myfielddict) )
    planid = myfielddict.get('',None)
    gwid = myfielddict.get('gw_id', None)
    modelid = myfielddict.get('model_id', None)

    # sql = "select plan_id from yw_project_boxing_info where gw_id='%s' and model_id='%s'" % (gwid, modelid)
    # mysql_conn=dbconnect(reconnect=True)
    # cursor=mysql_conn.cursor()
    # log.info(sql)
    # cursor.execute(sql)
    # row=cursor.fetchone()
    # if row == None or len(row) == 0:
    #     return "模块ID未装箱"
    #
    # db_planid=row[0]
    # if planid:
    #     if planid != db_planid:
    #         return "计划号不匹配"

    sql = "select plan_id from yw_project_snid_detail where gw_id='%s' and model_id='%s' and state='1'" % (gwid, modelid)
    mysql_conn=dbconnect(reconnect=True)
    cursor=mysql_conn.cursor()
    log.info(sql)
    cursor.execute(sql)
    row=cursor.fetchone()
    if row == None or len(row) == 0:
        return "模块ID与国网ID不匹配或不存在"

    # db_planid=row[0]
    # if '-' in db_planid:
    #     db_planid=db_planid.
    # if planid:
    #     if planid != db_planid:
    #         return "计划号不匹配"

    if not gwid.upper() == gwid:
        return "芯片ID存在小写情况"

    return '000000'


#处理文件
def filedue(fileid):
    log, fh = loger_init()
    log.info('----------------------fileid='+str(fileid)+' filedue_begin---------------------------')
    try:
        mysql_conn_mx = pymysql.Connect(
            host='192.168.2.174',
            port=3306,
            user='lqkj',
            passwd='LQkj666_2019',
            db='lqkj_db',
            charset='utf8',autocommit=0)
        cur = mysql_conn_mx.cursor()
    except Exception as e:
        log.info("MySql数据库连接失败!"+str(e) )
        mysql_conn_mx.close()
        return -1

    #先占用此条记录
    sql = "update irsadmin_db_unfile_info set resp_code='888888', resp_msg='%s', deal_num=0, repeat_num=0, error_num=0 " \
          "where file_id='%s' and resp_code='888887'" % ('开始处理',fileid)
    log.info(sql)
    if upd_file_info(sql) == False:
        #此记录已经被其它进程占用并处理
        mysql_conn.close()
        return

    #根据文件ID获取文件信息
    sql="select menu_id,file_path,file_name,sheet_name,file_rows,deal_num,deal_type from irsadmin_db_unfile_info where file_id='%s'" % fileid
    log.info(sql)
    cur.execute(sql)
    row=cur.fetchone()
    if not row:
        log.info("根据FILE_ID查询数据异常!" + str(fileid))
        mysql_conn.close()
        return -1
    fileinfo={
        "file_id": fileid,
        "menu_id": row[0],
        "file_path": row[1],
        "file_name": row[2],
        "sheet_name": row[3],
        "file_rows": row[4],
        "deal_num": row[5],
        "deal_type": row[6],
    }
    #根据menuid获取数据表等信息
    sql="select table_name from irsadmin_menu a, irsadmin_db_tran_reg b where a.APP_ID=b.APP_ID and a.MENU_ID='%s'" % fileinfo['menu_id']
    log.info(sql)
    cur.execute(sql)
    row = cur.fetchone()
    if not row:
        log.info("根据MENU_ID查询数据异常!" + str(fileinfo['menu_id']))
        mysql_conn.close()
        return -1
    tablename=row[0]

    # print('fileinfo=',fileinfo)
    # 根据文件ID获取文件导入配置信息
    sql = "select table_field,file_field from irsadmin_db_unfile_cfg " \
          "where file_id='%s' and FILE_FIELD is not null and FILE_FIELD!='' order by id asc" % fileid
    log.info(sql)
    cur.execute(sql)
    rows = cur.fetchall()
    if not rows:
        log.info("根据FILE_ID查询导入配置数据异常!" + str(fileid))
        mysql_conn.close()
        return -1

    fieldlist=""
    valueinfo=""
    valuelist=[]
    fielddict={}  #数据库字段，对应excel内容的数据字典

    for item in rows:

        if fieldlist == '':
            fieldlist = item[0]
            valueinfo = "'%s'"
        else:
            fieldlist = fieldlist+','+item[0]
            valueinfo = valueinfo+','+"'%s'"
        valuelist.append(item[1])
        fielddict[item[0]]=item[1]
        log.info(fieldlist)
    model_sql="insert into %s(%s) values(%s)" % (tablename, fieldlist, valueinfo)
    log.info(model_sql)

    #读文件
    file=fileinfo['file_path']+fileinfo['file_name']
    log.info("处理文件："+file)
    # if not os.path.exists(file):
    #     log.info("文件没有上传服务器或已删除!" + str(fileid))
    #     mysql_conn.close()
    #     return -1
    try:
        if file[-3:]=='xls':
            workbook = xlrd.open_workbook(filename=file, formatting_info=True)
            sheet1 = workbook.sheet_by_name(fileinfo['sheet_name'])
        elif file[-4:]=='xlsx':
            workbook = openpyxl.load_workbook(filename=file, read_only=True)
            sheet1 = workbook.get_sheet_by_name(fileinfo['sheet_name'])
    except Exception as ex:
        sql = "update irsadmin_db_unfile_info set resp_code='999999',resp_msg='%s' where file_id='%s'" \
              % (str(ex)[0:118].replace("'","\\'"), fileid)
        upd_file_info(sql)
        raise

    #开始处理excel文件的明细
    try:
        mysql_conn_mx.begin()
        deal_num = 1  # 处理记录数
        wrong_num = 0 #错误记录数
        for i in range(0, int( fileinfo['file_rows']) ):
            deal_num = deal_num +1 #从第二行开始
            values=[]
            for j in valuelist:
                if j and j!='':
                    if file[-3:] == 'xls':
                        values.append(sheet1.cell(deal_num-1, int(j)-1 ).value)
                    elif file[-4:] == 'xlsx':
                        values.append(sheet1.cell(deal_num, int(j) ).value)
                else:
                    values.append(None)
            myfielddict={}
            for item in fielddict:
                if j and j!='':
                    if file[-3:] == 'xls':
                        myfielddict[item] = sheet1.cell(deal_num - 1, int(fielddict[item]) - 1).value
                    elif file[-4:] == 'xlsx':
                        myfielddict[item] = sheet1.cell(deal_num, int(fielddict[item])).value
                else:
                    values.append(None)
            log.info( "menuid="+str( fileinfo['menu_id'] ) )
            if str( fileinfo['menu_id'] )=='191':#提供给客户芯片ID校验
                resp = fileup_proc_191( myfielddict )
                if resp!='000000':
                    wrong_num = wrong_num + 1
                    # 更新汇总表处理记录数
                    sql = "update irsadmin_db_unfile_info set deal_num='%s',error_num=error_num+1, resp_time=now()," \
                          "resp_msg=CONCAT(resp_msg, ';', '%s') " \
                          "where file_id='%s'" % (deal_num - 1, "第"+str(deal_num-1)+"行:"+resp, fileid)
                    upd_file_info(sql)
                    continue
            #直接执行插入操作
            # print( "model_sql value",values )
            sql =model_sql % (tuple(values))
            log.info(sql)
            cur.execute(sql)

            #更新汇总表处理记录数
            sql="update irsadmin_db_unfile_info set deal_num='%s', RESP_TIME=now() where file_id='%s'" % (deal_num-1, fileid)
            upd_file_info(sql)
    except Exception as ex:
        log.error('处理明细数据错误:' + str(ex), exc_info=True)
        mysql_conn_mx.rollback()
        mysql_conn_mx.commit()
        sql = "update irsadmin_db_unfile_info set resp_code='999999',resp_msg='%s' where file_id='%s' and resp_code='888888'" \
              % ("第"+str(deal_num-1)+"行:"+str(ex)[0:100].replace("'","\\'"), fileid)
        upd_file_info(sql)
        raise

    if wrong_num == 0:
        sql = "update irsadmin_db_unfile_info set resp_code='000000',resp_msg='批量文件处理成功', deal_num='%s' where file_id='%s' " \
              "and resp_code='888888'" % (deal_num-1, fileid)
        log.info(sql)
        if upd_file_info(sql) == False:
            # 更新文件登记表
            return
    mysql_conn_mx.commit()
    mysql_conn_mx.close()

    log.info('----------------------fileid='+str(fileid)+' filedue_end---------------------------')

#更新文件汇总表记录数
def upd_file_info( sql ):
    mysql_conn = dbconnect(reconnect=True)
    if mysql_conn == False:
        log.info('连接数据库失败')
        raise
    cur = mysql_conn.cursor()
    log.info('更新文件汇总表SQL:'+str(sql))
    cur.execute(sql)
    mysql_conn.commit()
    if cur.rowcount == 0:
        return False
    else:
        return True

#检索是否有需要处理的时间
def fileselect():
    log, fh = loger_init()
    mysql_conn = dbconnect(reconnect=True)
    if mysql_conn == False:
        log.info('连接数据库失败')
        raise
    cur = mysql_conn.cursor()
    sql="select file_id from irsadmin_db_unfile_info where resp_code='888887'"
    cur.execute(sql)
    rows=cur.fetchall()
    FileId_List=[]
    for item in rows:
        FileId_List.append(item[0])
    cur.close()
    mysql_conn.close()

    # log.info('----------------------获取待处理的记录数:'+str(len(FileId_List))+'-----------------------')
    for item in FileId_List:
        filedue(item)

#持续运行
if __name__ == '__main__':
    while True:
        log, fh = loger_init()
        try:
            fileselect()
        except Exception as ex:
            log.error('程序运行错误:'+str(ex), exc_info=True)
            # print(ex)
        log.removeHandler(fh)
        time.sleep(3)

