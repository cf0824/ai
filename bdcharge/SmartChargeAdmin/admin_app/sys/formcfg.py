import sys
from django.shortcuts import render,redirect,HttpResponse
from django.db import connection, transaction
import json
from admin_app.sys import public
import datetime
import re
import openpyxl
import base64
from io import BytesIO

#配置操作主流程
@transaction.atomic()
def Main_Proc( request ):
    public.respcode, public.respmsg = "999998", "交易开始处理!"
    log = public.logger
    sid = transaction.savepoint()
    func_name=public.tran_type+'(request)'
    if globals().get(public.tran_type):
        log.info('---[%s]-begin---' % (public.tran_type), extra={'ptlsh': public.req_seq})
        public.respinfo = eval(func_name)
        log.info('---[%s]-end----' % (public.tran_type), extra={'ptlsh': public.req_seq})
    else:
        public.respcode, public.respmsg = "100002", "trantype error!"
        public.respinfo = HttpResponse( public.setrespinfo() )
    if public.respcode=="000000":
        # 提交事务
        transaction.savepoint_commit(sid)
    # else:
    #     # 回滚事务
    #     transaction.savepoint_rollback(sid)
    return public.respinfo


# 根据formcfg获取table组件字段明细,并插入字段明细表，供后续权限使用。
def FormTableFieldIns(formid, formcfg):
    log = public.logger
    log.info(f'更新formcfg字段明细')
    cur = connection.cursor()  # 创建游标
    sql = "delete from sys_form_table_field where form_id=%s"
    cur.execute(sql, formid)

    log.info(formcfg)
    for comp in formcfg:

        if not isinstance(comp, dict):

            continue
        if comp.get('type') != 'table':

            continue
        table_comp_id = comp.get('id')
        table_comp_attrs = comp.get('attrs')
        log.info('table_comp_id=%s' % table_comp_id)
        # print('table_comp_attrs=', table_comp_attrs)
        head_button_diy = table_comp_attrs.get('head_button_diy')
        log.info('head_button_diy=%s' % str(head_button_diy))
        # head_button_diy = "head_button_diy":[{"button_title":"按钮1","button_size":"mini","button_color":"primary","button_icon":"","button_plain":0,"button_form_id":"1302","button_api":"","button_tran_type":"","open_url":"","close_form":false,"show_vote_key":"","show_vote_value":"","ack_msg":""}]
        for button_diy in head_button_diy:

            button_id = button_diy.get('button_title')
            sql = "insert into sys_form_table_field(form_id, comp_id, field_id, field_type, field_name, show_able, dis_able) " \
                  "values('%s', '%s', '%s',' %s','%s', '%s', '%s' )" \
                  % (formid, table_comp_id, button_id, 'diybutton', button_id, 'Y', 'N')
            log.info(f'sql={sql}')
            cur.execute(sql)

        buttonSwitch = table_comp_attrs.get('buttonSwitch')
        log.info('buttonSwitch=%s' % str(buttonSwitch))
        for button_id in buttonSwitch:

            sql = "insert into sys_form_table_field(form_id, comp_id, field_id, field_type, field_name, show_able, dis_able) " \
                  "values('%s', '%s', '%s', '%s', '%s', '%s', '%s' )" \
                  % (formid, table_comp_id, button_id, 'button', button_id, 'Y', 'N')
            log.info(f'sql={sql}')
            cur.execute(sql)

        head = table_comp_attrs.get('head')
        log.info('head=%s' % str(head))
        for fieldinfo in head:

            field_id = fieldinfo.get('name')
            # field_type = fieldinfo.get('comp')
            field_type = 'table_head'
            field_name = fieldinfo.get('label')
            field_show_able = public.True2y(fieldinfo.get('show'))
            sql = "insert into sys_form_table_field(form_id, comp_id, field_id, field_type, field_name, show_able, dis_able) " \
                  "values('%s', '%s', '%s', '%s', '%s', '%s', '%s' )" \
                  % (formid, table_comp_id, field_id, field_type, field_name, field_show_able, 'N')
            log.info(f'sql={sql}')
            cur.execute(sql)

    cur.close()


# 根据formcfg获取字段明细,并插入字段明细表
def FormcfgFieldListIns(formid,formcfg):
    # log = public.logger
    # 递归获取表单所有组件
    def GetCompFormLayout(formcfg):
        comp_list = []
        for comp_item in formcfg:
            # print(comp_item)
            itemdict = {}
            if comp_item.get('children'):
                comp_list = comp_list + GetCompFormLayout(comp_item['children'])
                continue
            elif comp_item['type'] == 'null':
                continue
            elif comp_item['type'] == 'button': #按钮的话，取trantype做为field_id
                itemdict['field_id'] = comp_item['attrs']['tran_type']
            else:
                itemdict['field_id'] = comp_item['attrs']['variable']

            if '显示' in comp_item['attrs']['power']:
                itemdict['show_able'] = 'Y'
            else:
                itemdict['show_able'] = 'N'
            if '置灰' in comp_item['attrs']['power']:
                itemdict['dis_able'] = 'Y'
            else:
                itemdict['dis_able'] = 'N'
            itemdict['comp_id'] = comp_item['id']
            itemdict['comp_type'] = comp_item['type']
            itemdict['field_name'] = comp_item['attrs'].get('label')
            comp_list.append(itemdict)
        return comp_list

    comp_list = GetCompFormLayout(formcfg)
    cur = connection.cursor()  # 创建游标
    sql = "delete from sys_form_cfg_fieldlist where form_id=%s"
    cur.execute(sql, formid)
    for itm in comp_list:
        sql = "insert into sys_form_cfg_fieldlist(form_id, comp_id, comp_type, field_id, field_name, show_able, dis_able) " \
              "values(%s, %s, %s, %s, %s, %s, %s )"
        # log.info(sql % (formid, itm.get('comp_id'), itm.get('comp_type'), itm.get('field_id'),
        #                 itm.get('field_name'), itm.get('show_able','N'), itm.get('dis_able','N') )
        #          , extra={'ptlsh': public.req_seq})
        cur.execute(sql, (formid, itm.get('comp_id'), itm.get('comp_type'), itm.get('field_id'),
                          itm.get('field_name'), itm.get('show_able','N'), itm.get('dis_able','N') ) )
    cur.close()

#表单配置新增，主要是获取form_id
def form_cfg_create( request ):
    log = public.logger
    body = public.req_body  # 请求报文体
    form_id = body.get('form_id')
    form_name = body.get('form_name')
    form_show_tran_type = body.get('form_show_tran_type')
    form_show_api = body.get('form_show_api')
    form_cfg = body.get('form_cfg')
    # form_var = body.get('form_var')
    form_var = "{}" #创建获取ID，变量内容不需要
    form_attr = "{}" #表单属性

    # 递归获取表单按钮组件的SQL
    def GetFormButtonSql(Layout_list):
        comp_list = {}
        for comp_item in Layout_list:
            if comp_item.get('children'):
                # comp_list = comp_list + GetFormButtonSql(comp_item['children'])
                comp_list.update(GetFormButtonSql(comp_item['children']))
                continue
            elif comp_item['type'] == 'button':
                # button_key=comp_item['attrs']['tran_type']+'_'+comp_item['id']
                tran_append_sql = comp_item['attrs'].get('tran_append_sql')
                comp_list[comp_item['id']]=tran_append_sql
            else:
                continue
        return comp_list

    try:
        form_cfg_json = json.loads(body.get('form_cfg'))
        form_sql = GetFormButtonSql( form_cfg_json )
        if form_sql:
            form_sql = json.dumps(form_sql)
        # log.info('form_sql:'+str(form_sql), extra={'ptlsh': public.req_seq})
        cur = connection.cursor()  # 创建游标
        if form_id:
            sql = "select CONCAT(form_name,'_copy'), form_cfg, form_var, form_sql, form_show_tran_type, form_show_api, form_attr " \
                  "from sys_form_cfg_info where form_id=%s"
            cur.execute(sql,form_id)
            row = cur.fetchone()
            if not row:
                public.respcode, public.respmsg = "100010", "表单信息不存在!"
                public.respinfo = HttpResponse(public.setrespinfo())
                return public.respinfo
            if not form_name:
                form_name=row[0]
            form_cfg = row[1]
            form_var = row[2]
            form_sql = row[3]
            form_show_tran_type = row[4]
            form_show_api = row[5]
            form_attr = row[6]
        else:
            form_show_tran_type = 'form_cfg_show'
            form_show_api = '/interface/sys/formcfg'
            form_cfg = "[]"
            form_var = "{}"
            form_attr = "{}"

        if form_sql:
            sql="insert into sys_form_cfg_info(form_name, form_show_tran_type, form_show_api, user_id, form_cfg, form_var, form_attr, form_sql) " \
                "values(%s, %s, %s, %s, %s, %s, %s, %s)"
            cur.execute(sql, ( form_name,form_show_tran_type,form_show_api, public.user_id, form_cfg, form_var, form_attr, form_sql) )
        else:
            sql="insert into sys_form_cfg_info(form_name, form_show_tran_type, form_show_api,  user_id, form_cfg, form_var, form_attr ) " \
                "values(%s, %s, %s, %s, %s, %s, %s)"
            cur.execute(sql, ( form_name, form_show_tran_type,form_show_api, public.user_id, form_cfg, form_var, form_attr ) )
        #查询刚刚插入的ID
        cur.execute("SELECT LAST_INSERT_ID()") #获取自增字段刚刚插入的ID
        row=cur.fetchone()
        if row:
            form_id=row[0]
            log.info('FormID生成，自增字段ID:%s' % str(form_id), extra={'ptlsh': public.req_seq})
        cur.close() #关闭游标

        #登记FORM字段属性表
        FormcfgFieldListIns(form_id, form_cfg_json)

    except Exception as ex:
        log.error("登记配置信息失败!"+str(ex), exc_info=True, extra={'ptlsh':public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        public.respcode, public.respmsg = "100010", "登记配置信息失败!"
        public.respinfo = HttpResponse( public.setrespinfo() )
        return public.respinfo

    public.respcode, public.respmsg = "000000", "新增配置成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": {
            "form_id":form_id
        }
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo

#表单配置修改
def form_cfg_update( request ):
    log = public.logger
    body = public.req_body  # 请求报文体
    form_id = body.get('form_id')
    form_cfg = body.get('form_cfg')
    form_var = body.get('form_var')
    form_attr = body.get('form_attr')

    form_show_tran_type = body.get('form_show_tran_type')
    form_show_api = body.get('form_show_api')

    # 递归获取表单按钮组件的SQL
    def GetFormButtonSql(Layout_list):
        comp_list = {}
        for comp_item in Layout_list:
            if comp_item.get('children'):
                # comp_list = comp_list + GetFormButtonSql(comp_item['children'])
                comp_list.update( GetFormButtonSql(comp_item['children']))
                continue
            elif comp_item['type'] == 'button':
                # button_key=comp_item['attrs']['tran_type']+'_'+comp_item['id']
                tran_append_sql = comp_item['attrs'].get('tran_append_sql')
                comp_list[comp_item['id']]=tran_append_sql
            else:
                continue
        return comp_list

    if not form_id:
        public.respcode, public.respmsg = "100021", "form_id不可为空!"
        public.respinfo =  HttpResponse( public.setrespinfo() )
        return public.respinfo

    try:
        if form_cfg:
            form_cfg_json = json.loads(form_cfg)
            form_sql = GetFormButtonSql( form_cfg_json )
        else:
            form_cfg_json = ''
            form_sql = {}
        form_sql = json.dumps(form_sql)

        cur = connection.cursor()  # 创建游标
        if body.get('form_var'):
            sql="update sys_form_cfg_info set form_attr=%s, form_var=%s, update_user=%s,update_date=%s where form_id=%s"
            cur.execute(sql, ( form_attr, form_var, public.user_id, datetime.datetime.now(), form_id) )
        elif body.get('form_cfg'):
            sql = "update sys_form_cfg_info set form_attr=%s, form_name=%s, form_show_tran_type=%s, form_show_api=%s,  " \
                  "form_cfg=%s, form_sql=%s, update_user=%s,update_date=%s where form_id=%s"
            cur.execute(sql, ( form_attr, body.get('form_name'), form_show_tran_type, form_show_api, form_cfg, form_sql,
                               public.user_id, datetime.datetime.now(), form_id) )
        else:
            sql = "update sys_form_cfg_info set form_attr=%s, form_name=%s, form_show_tran_type=%s, form_show_api=%s, " \
                  "form_cfg=%s, form_var=%s, form_sql=%s, update_user=%s,update_date=%s where form_id=%s"
            cur.execute(sql, ( form_attr, body.get('form_name'), form_show_tran_type, form_show_api, form_cfg, form_var, form_sql,
                               public.user_id, datetime.datetime.now(), form_id))
        cur.close() #关闭游标

        if form_cfg_json:
            # 登记FORM字段属性表
            FormcfgFieldListIns(form_id, form_cfg_json)
            # 登记TABLE组件按钮和字段列表
            FormTableFieldIns(form_id, form_cfg_json)

    except Exception as ex:
        log.error("更新配置信息失败!"+str(ex), exc_info=True, extra={'ptlsh':public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        public.respcode, public.respmsg = "100011", "更新配置信息失败!"
        public.respinfo =  HttpResponse( public.setrespinfo() )
        return public.respinfo

    public.respcode, public.respmsg = "000000", "更新配置成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": {
            "form_id":form_id
        }
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo


# 表单配置信息查询
def form_cfg_select(request):
    log = public.logger
    body = public.req_body
    form_id = body.get('form_id', '')
    editor = body.get('editor', None)

    try:
        cur = connection.cursor()  # 创建游标

        if not form_id: #没有formid ,是菜单直接进入的交易。根据menuid取formid
            sql = "select app_id from sys_menu where menu_id=%s "
            cur.execute(sql, (public.menu_id))
            row = cur.fetchone()
            if not row:
                cur.close()
                log.info("对应的APPID不存在!", extra={'ptlsh': public.req_seq})
                public.respcode, public.respmsg = "100111", "对应的APPID不存在!"
                public.respinfo = HttpResponse(public.setrespinfo())
                return public.respinfo

            form_id = row[0]

        sql="select form_name,form_show_tran_type, form_show_api, form_cfg,form_var, form_attr " \
            "from sys_form_cfg_info where form_id=%s"
        cur.execute(sql, form_id)
        row = cur.fetchone()
        if row:
            form_name = row[0]
            form_show_tran_type = row[1]
            form_show_api = row[2]
            form_cfg = row[3]
            form_cfg = json.loads(form_cfg, encoding='utf-8')
            form_var = row[4]
            if form_var:
                form_var = json.loads(form_var, encoding='utf-8')
            else:
                form_var={}
            form_attr = row[5]
            if form_attr:
                form_attr = json.loads(form_attr, encoding='utf-8')
            else:
                form_attr={}
        else:
            form_name = ''
            form_show_tran_type = 'form_cfg_show'
            form_show_api = '/interface/sys/formcfg'
            form_cfg = []
            form_var = {}
            form_attr = {}

        if not editor:  # 非编辑模式，返回菜单字段权限
            # 获取表单组件权限
            sql = "select distinct a.field_id, a.field_type, a.field_name, show_able, dis_able " \
                  "from sys_role_purv_table_field a, sys_user_role b " \
                  "where a.ROLE_ID=b.ROLE_ID and a.FORM_ID='%s' and b.USER_ID='%s' " % (form_id, public.user_id)
            cur.execute(sql)
            rows = cur.fetchall()
            table_auth_flag = False  # 是否处理表格按钮字段权限
            if rows:
                table_auth_flag = True

            # 当前用户是否具体root权限
            root_flag = False
            sql = "select role_id from sys_user_role where role_id='root' and user_id='%s'" % public.user_id
            cur.execute(sql)
            row = cur.fetchone()
            if row:
                root_flag = True

            # 判断字段是否显示
            for comp_item in form_cfg:
                # 非编辑模式，隐藏SQL
                if comp_item['attrs'].get('tran_append_sql'):
                    comp_item['attrs']['tran_append_sql'] = ''  # 按钮中的交易附加SQL
                if comp_item['attrs'].get('render') and 'GetValueFromSQL' in comp_item['attrs'].get('render'):
                    comp_item['attrs']['render'] = ''  # 组件中的渲染规则
                if comp_item['attrs'].get('options_render') and 'GetValueFromSQL' in comp_item['attrs'].get('options_render'):
                    comp_item['attrs']['options_render'] = ''  # 组件中的选项渲染规则
                if comp_item['attrs'].get('table_render'):
                    comp_item['attrs']['table_render'] = ''  # 表格组件中的获取数据sql
                if comp_item['attrs'].get('data_sql'):
                    comp_item['attrs']['data_sql'] = ''  # 表格组件中的获取数据sql
                if comp_item['attrs'].get('delete_append_sql'):
                    comp_item['attrs']['delete_append_sql'] = ''  # 表格组件中的删除数据sql

                # 表单组件按钮和字段的权限
                if comp_item.get('type') == 'table' and table_auth_flag and not root_flag:  # table按钮权限控制
                    table_comp_attrs = comp_item.get('attrs')
                    head_button_diy = table_comp_attrs.get('head_button_diy')
                    # log.info("head_button_diy:" + str(head_button_diy))
                    # buttonSwitch = table_comp_attrs.get('buttonSwitch')
                    # log.info("buttonSwitch:" + str(buttonSwitch))
                    db_diybutton_list = []
                    sql = "select distinct a.field_id " \
                          "from sys_role_purv_table_field a, sys_user_role b " \
                          "where a.ROLE_ID=b.ROLE_ID and field_type='diybutton' and show_able='Y' " \
                          "and a.FORM_ID='%s' and b.USER_ID='%s' " % (form_id, public.user_id)
                    log.info("sql1:" + sql)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    for item in rows:
                        db_diybutton_list.append(item[0])
                    log.info("db_diybutton_list:" + str(db_diybutton_list))
                    new_head_button_diy = []
                    for item in head_button_diy:
                        if item.get('button_title') in db_diybutton_list:
                            new_head_button_diy.append(item)
                    comp_item['attrs']['head_button_diy'] = new_head_button_diy
                    log.info("head_button_diy:" + str(comp_item['attrs']['head_button_diy']))

                    comp_item['attrs']['buttonSwitch'] = []
                    sql = "select distinct a.field_id " \
                          "from sys_role_purv_table_field a, sys_user_role b " \
                          "where a.ROLE_ID=b.ROLE_ID and field_type='button' and show_able='Y' " \
                          "and a.FORM_ID='%s' and b.USER_ID='%s' " % (form_id, public.user_id)
                    log.info("sql2:" + sql)
                    cur.execute(sql)
                    rows = cur.fetchall()
                    for item in rows:
                        comp_item['attrs']['buttonSwitch'].append(item[0])
                    log.info("buttonSwitch:" + str(comp_item['attrs']['buttonSwitch']))
                #
                # if comp_item.get('children'):
                #     comp_item['children'] = DueCompFormLayout(comp_item['children'])
                # elif comp_item['type'] == 'null':
                #     pass
                # else:
                #     if comp_item['type'] == 'button':  # 按钮的话，取trantype做为field_id
                #         field_id = comp_item['attrs']['tran_type']
                #     else:
                #         field_id = comp_item['attrs'].get('variable')
                #     if field_id:
                #         field_id=field_id.upper()
                #         log.info('role_purv_body2=' + field_id+'-show_able-Y', extra={'ptlsh': public.req_seq})
                #
                # form_cfg.append(comp_item)
                # log.info('form_cfg:' + json.dumps(form_cfg, ensure_ascii=False),  extra={'ptlsh': public.req_seq})

        # 最终返回数据需要转为字符串
        if form_cfg:
            form_cfg = json.dumps(form_cfg )  # 将配置转换为字符串
        else:
            form_cfg = "[]"
        if form_var:
            form_var = json.dumps(form_var )  # 将配置转换为字符串
        else:
            form_var = "{}"
        if form_attr:
            form_attr = json.dumps(form_attr)  # 将配置转换为字符串
        else:
            form_attr = "{}"

        cur.close() #关闭游标
    except Exception as ex:
        log.error("表单配置信息查询失败!"+str(ex), exc_info=True, extra={'ptlsh':public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        cur.close()  # 关闭游标
        public.respcode, public.respmsg = "100110", "表单配置信息查询失败!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    public.respcode, public.respmsg = "000000", "表单配置信息查询成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": {
            "form_id":form_id,
            "form_name":form_name,
            "form_show_tran_type": form_show_tran_type,
            "form_show_api": form_show_api,
            "form_cfg":form_cfg,
            "form_var":form_var,
            "form_attr":form_attr,
        }
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo


#表单渲染显示
def form_cfg_show( request ):
    log = public.logger
    body = public.req_body
    form_id = body.get('form_id', '')
    form_data = body.get('form_data', '')   # 可能为空
    extra_data = body.get('extra')

    if not form_data:
        form_data = {}
    try:
        cur = connection.cursor()  # 创建游标
        if not form_id: #没有formid ,是菜单直接进入的交易。根据menuid取formid
            sql = "select app_id from sys_menu where menu_id=%s "
            cur.execute(sql, (public.menu_id))
            row = cur.fetchone()
            if not row:
                cur.close()
                log.info("对应的APPID不存在!", extra={'ptlsh': public.req_seq})
                public.respcode, public.respmsg = "100111", "对应的APPID不存在!"
                public.respinfo = HttpResponse(public.setrespinfo())
                return public.respinfo

            form_id = row[0]

        sql = "select form_cfg, form_var from sys_form_cfg_info where form_id=%s"
        # log.info("查询表单配置信息:" + sql % form_id, extra={'ptlsh': public.req_seq})
        cur.execute(sql, (form_id) )
        row = cur.fetchone()
        if row:
            form_cfg = row[0]
            form_var = row[1]
        else:
            cur.close()
            log.info("对应的FORMID不存在!", extra={'ptlsh': public.req_seq})
            public.respcode, public.respmsg = "100111", "对应的FORMID不存在!"
            public.respinfo = HttpResponse(public.setrespinfo())
            return public.respinfo

        cur.close()

    except Exception as ex:
        log.error("表单配置信息查询失败!"+str(ex), exc_info=True, extra={'ptlsh':public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        public.respcode, public.respmsg = "100110", "表单配置信息查询失败!"
        public.respinfo =  HttpResponse( public.setrespinfo() )
        return public.respinfo

    #对关键字进行处理
    def keywords( keyword ):
        new_value=keyword
        if keyword == '${USER_ID}':
            new_value=public.user_id
        elif keyword == '${USER_NAME}':
            sql="select user_name from sys_user where user_id='%s'" % public.user_id
            cur.execute(sql)
            row=cur.fetchone()
            new_value=row[0]
        elif keyword == '${TRAN_DATE}':
            new_value = datetime.datetime.now().strftime('%Y-%m-%d')
        elif keyword == '${TRAN_TIME}':
            new_value = datetime.datetime.now().strftime('%H:%M:%S')
        elif keyword == '${TRAN_DATETIME}':
            new_value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elif keyword == '${YYYY}':  #当前年份
            new_value = datetime.datetime.now().strftime('%Y')
        elif keyword == '${WEEK}':  #当年第几周
            new_value = datetime.datetime.now().strftime('%W')
        # log.info('keyword=' + str(keyword), extra={'ptlsh': public.req_seq})
        # log.info('new_value='+str(new_value), extra={'ptlsh':public.req_seq})
        return new_value

    # 递归获取表单所有组件
    def GetCompFormLayout(Layout_list):
        comp_list = []
        for comp_item in Layout_list:
            if comp_item.get('children'):
                comp_list = comp_list + GetCompFormLayout(comp_item['children'])
                continue
            elif comp_item['type'] == 'null':
                continue
            else:
                comp_list.append(comp_item)
        return comp_list

    # 处理配置的sql变量
    def GetValueFromSQL(render, render_type):
        # log.info('render='+str(render), extra={'ptlsh':public.req_seq})
        pattern = re.compile("GetValueFromSQL{(.*)}")
        log.info('render=' + str(render), extra={'ptlsh': public.req_seq})
        sql = pattern.findall(render)[0]
        log.info('render sql=' + str(sql), extra={'ptlsh': public.req_seq})
        pattern = re.compile("\$\[(.*?)\]")
        sqlvar = pattern.findall(sql)
        for sqlitm in sqlvar:
            old = "$[" + sqlitm + "]"
            if sqlitm in form_data.keys():
                new = "'" + str(form_data.get(sqlitm)) + "'"
            else:
                new = "''"
            sql = sql.replace(old, new)
        log.info('real sql=' + str(sql), extra={'ptlsh': public.req_seq})
        sql = public.SqlKeywordConver(sql, form_var_dict)
        log.info('finally sql=' + str(sql), extra={'ptlsh': public.req_seq})
        cur.execute(sql)
        if render_type == 'options_render':  # 列表
            reslist = []
            rows = cur.fetchall()
            for resitm in rows:
                if len(resitm) == 2:
                    kv={}
                    kv['key'] = resitm[0]
                    kv['value'] = resitm[1]
                    reslist.append(kv)
                else:
                    reslist.append(resitm)
            return reslist
        elif render_type == 'table_render':  # 表络
            reslist = []
            rows = cur.fetchall()
            for resitm in rows:
                kv = {}
                i = 0
                for tabitem in compitem['attrs'].get('head'):
                    kv[tabitem.get('name')] = resitm[i]
                    i = i+1
                reslist.append(kv)
            return reslist
        elif render_type == 'list_render':  # LIST列表
            reslist = []
            rows = cur.fetchone()
            for resitm in rows:
                reslist.append(resitm)
            return reslist

        elif render_type == 'render':  # 数据
            reslist = ""
            rows = cur.fetchone()
            if rows: #查到数据了。
                reslist = rows[0]
            return reslist
        else:
            return ""

    #开始渲染表单
    try:
        cur = connection.cursor()  # 创建游标

        form_cfg = json.loads(form_cfg)
        form_var = json.loads(form_var)
        comp_list = GetCompFormLayout(form_cfg)  # 组件列表
        form_var_dict={}
        for compitem in comp_list:
            comptype=compitem.get('type')
            if comptype in ('button'):
                continue

            options_render = compitem['attrs'].get('options_render')
            options = compitem['attrs'].get('options')
            variable = compitem['attrs'].get('variable')
            if '.' in variable or  '[' in variable or  ']' in variable:  #前端解析多级结构
                continue

            render = compitem['attrs'].get('render')
            # log.info( "compitem="+str(compitem), extra={'ptlsh': public.req_seq})
            if comptype in ['date', 'datetime', 'input']:  # 日期时间，使用eval也不会转换错
                if form_data and form_data.get(variable):
                    form_var_dict[variable] = form_data.get(variable)
                elif render:  #值渲染
                    if 'GetValueFromSQL' in render:
                        try:
                            form_var_dict[variable] = GetValueFromSQL(render, 'render')
                        except Exception as ex:
                            log.error('表单配置SQL错误:' + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
                            public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
                            public.respcode, public.respmsg = "100021", "表单配置项[%s]SQL错误!" % compitem['attrs'].get('label')
                            public.respinfo = HttpResponse(public.setrespinfo())
                            return public.respinfo
                    else:
                        if variable:
                            if form_data.get(variable):
                                form_var_dict[variable] = form_data.get(variable)
                            else:
                                form_var_dict[variable] = render
                else:
                    form_var_dict[variable] = ''
            elif comptype in ['select', 'radio']:  #下拉、选项字段的下拉内容渲染
                if options_render: #选项渲染, 用来放SQL语句
                    if 'GetValueFromSQL' in options_render:
                        try:
                            form_var_dict[options] = GetValueFromSQL(options_render, 'options_render')
                        except Exception as ex:
                            log.error('表单配置SQL错误:' + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
                            public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
                            public.respcode, public.respmsg = "100021", "表单配置项[%s]SQL错误!" % compitem['attrs'].get('label')
                            public.respinfo = HttpResponse(public.setrespinfo())
                            return public.respinfo
                    else:
                        form_var_dict[options] = options_render
                else: #没有SQL语句，直接取options的变量值
                    if options:
                        form_var_dict[options] = form_var.get(options)

                if variable: #对绑定变量初始化
                    if form_data.get(variable):
                        form_var_dict[variable] = form_data.get(variable)
                    else:
                        form_var_dict[variable] = render
            elif comptype in ['checkbox']:  # 多选框下拉、选项字段的下拉内容渲染
                log.info(f'compitem={compitem}, variable={variable}')

                options_cfg = compitem['attrs'].get('options_cfg')
                op_data_type = options_cfg.get('data_type')
                op_data_sql = options_cfg.get('data_sql')
                op_kv_tran = options_cfg.get('kv_tran')
                if op_data_type == 1: # 从SQL语句中获取变量信息
                    if op_data_sql:
                        try:
                            op_data_sql = public.SqlKeywordConver(op_data_sql, form_var_dict)
                            # log.info("op_data_sql=" + str(op_data_sql), extra={'ptlsh': public.req_seq})
                            cur.execute(op_data_sql)
                            rows = cur.fetchall()
                            reslist=[]
                            for resitm in rows:
                                if len(resitm) == 2:
                                    kv = {}
                                    kv['key'] = resitm[0]
                                    kv['value'] = resitm[1]
                                    reslist.append(kv)
                                else:
                                    reslist.append(resitm)
                            form_var_dict[options] = reslist
                        except Exception as ex:
                            log.error('表单配置SQL错误:' + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
                            public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
                            public.respcode, public.respmsg = "100021", "表单配置项[%s]SQL错误!" % compitem['attrs'].get('label')
                            public.respinfo = HttpResponse(public.setrespinfo())
                            return public.respinfo
                    else:
                        form_var_dict[options] = op_kv_tran
                else: # 没有SQL语句，直接取options的变量值
                    if options:
                        form_var_dict[options] = op_kv_tran

                if variable: # 对绑定变量初始化
                    if form_data.get(variable):
                        form_var_dict[variable] = form_data.get(variable)
                    elif render:
                        form_var_dict[variable] = GetValueFromSQL(render, 'list_render')
                    else:
                        form_var_dict[variable] = []
                # log.info('%s %s %s' % (variable, type(form_var_dict.get(variable)), form_data.get(variable)))

            elif comptype in ['table']:  # 表格内容渲染
                # tablehead = compitem['attrs']['head'][0]
                # print(type(tablehead), tablehead.keys())
                # if 'table' in tablehead.keys(): #增删改查配置的表格
                #     print('istable==', tablehead.get('table'))
                # el
                if render:  # 选项渲染, 用来放SQL语句
                    if 'GetValueFromSQL' in render:
                        try:
                            form_var_dict[variable] = GetValueFromSQL(render, 'table_render')
                        except Exception as ex:
                            log.error('表单配置SQL错误:' + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
                            public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
                            public.respcode, public.respmsg = "100021", "表单配置项[%s]SQL错误!" % compitem['attrs'].get('variable')
                            public.respinfo = HttpResponse(public.setrespinfo())
                            return public.respinfo
                    else:
                        form_var_dict[variable] = render
                else:  # 没有SQL语句，直接取options的变量值
                    if variable:
                        form_var_dict[variable] = form_var.get(variable)
            elif comptype in ['transfer']: #穿梭框
                if render and 'FUNC{' in render:  # 调用指定函数
                    from admin_app.tranapp import transfer
                    func_name = "transfer."+render[5:-1]+"( request )"
                    log.info("开始执行穿梭框自定义函数："+str(func_name), extra={'ptlsh': public.req_seq})
                    dataVariable = compitem['attrs'].get('dataVariable')
                    form_var_dict[variable], form_var_dict[dataVariable] = eval(func_name)
                else:
                    form_var_dict[variable] = ''
            elif comptype in ['tree']: #树结构
                if render and 'FUNC{' in render:  # 调用指定函数
                    from admin_app.tranapp import tree
                    func_name = "tree."+render[5:-1]+"( request )"
                    log.info("开始执行树结构自定义函数："+str(func_name), extra={'ptlsh': public.req_seq})
                    selectedVariable = compitem['attrs'].get('selectedVariable')
                    form_var_dict[variable], form_var_dict[selectedVariable] = eval(func_name)
                else:
                    form_var_dict[variable] = ''
            elif comptype in ['json_editor']: #JSON报文编辑框
                if form_data and form_data.get(variable):
                    form_var_dict[variable] = form_data.get(variable)
                elif render:  #值渲染
                    if 'GetValueFromSQL' in render:
                        try:
                            form_var_dict[variable] = GetValueFromSQL(render, 'render')
                        except Exception as ex:
                            log.error('表单配置SQL错误:' + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
                            public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
                            public.respcode, public.respmsg = "100021", "表单配置项[%s]SQL错误!" % compitem['attrs'].get('label')
                            public.respinfo = HttpResponse(public.setrespinfo())
                            return public.respinfo
                    else:
                        if variable:
                            if form_data.get(variable):
                                form_var_dict[variable] = form_data.get(variable)
                            else:
                                form_var_dict[variable] = render
                else:
                    form_var_dict[variable] = ''
            else: # 其它组件的变量初始化
                if render:  # 值渲染
                    if 'GetValueFromSQL' in render:
                        try:
                            form_var_dict[variable] = GetValueFromSQL(render, 'render')
                        except Exception as ex:
                            log.error('表单配置SQL错误:' + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
                            public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
                            public.respcode, public.respmsg = "100021", "表单配置项[%s]SQL错误!" % compitem['attrs'].get('label')
                            public.respinfo = HttpResponse(public.setrespinfo())
                            return public.respinfo
                    else:
                        if variable:
                            if form_data.get(variable):
                                try:
                                    form_var_dict[variable] = eval(form_data.get(variable))
                                except:
                                    form_var_dict[variable] = form_data.get(variable)
                            else:
                                form_var_dict[variable] = render
                else:
                    if variable:
                        if comptype in ['file_upload', 'img_upload']:  # 文件、图片上传。返回list
                            if form_data and form_data.get(variable):
                                try:
                                    form_var_dict[variable] = eval(form_data.get(variable))
                                except:
                                    form_var_dict[variable] = form_data.get(variable)
                            else:
                                form_var_dict[variable] = ''
                        else:
                            # log.info('variable--:'+str(variable))
                            if form_data and form_data.get(variable):
                                try:
                                    # modify by litz, 20221209
                                    # form_var_dict[variable] = eval(form_data.get(variable))
                                    form_var_dict[variable] = form_data.get(variable)
                                    log.info('%s %s %s' % (variable, type(form_var_dict.get(variable)), form_data.get(variable)))
                                except:
                                    form_var_dict[variable] = form_data.get(variable)
                            else:
                                form_var_dict[variable] = form_var.get(variable)
                            # log.info('variable--b - e :' + str(type(form_data.get(variable))) +'---'
                            # + str(type(form_var_dict.get(variable))) )

            # 关键字段值转换
            if form_var_dict.get(variable):
                form_var_dict[variable] = keywords(form_var_dict[variable])

            # log.info('%s %s %s' % (variable, type(form_var_dict.get(variable)), form_data.get(variable)))
        # 把form_var中独有的变量，也加入到返回列表中。
        for item in form_var:
            if not form_var_dict.get(item):
                form_var_dict[item] = form_var[item]

        # 把form_data中独有的变量，也加入到返回列表中。
        for item in form_data:
            print(item, type(form_data[item]))
            if not form_var_dict.get(item):
                form_var_dict[item] = form_data[item]

        # del form_var_dict['selected[0].name']
        # del form_var_dict['selected[0].sex']
        # del form_var_dict['selected[0].age']

        cur.close()

    except Exception as ex:
        log.error("表单配置[%s]渲染失败!" % variable + str(ex), exc_info=True, extra={'ptlsh':public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        public.respcode, public.respmsg = "100011", "表单配置渲染失败!"+str(variable)
        public.respinfo =  HttpResponse( public.setrespinfo() )
        return public.respinfo

    cur.close()  # 关闭游标
    public.respcode, public.respmsg = "000000", "表单配置信息查询成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": {
            "form_id":form_id,
            # "form_cfg":form_cfg,
            "form_var":form_var_dict,
        }
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo

#表单配置查询，主要是获取form_id列表
def form_cfg_list( request ):
    log = public.logger
    body = public.req_body
    form_id = body.get('form_id', '')
    try:
        cur = connection.cursor()  # 创建游标
        sql = "select form_id, form_name, update_date from sys_form_cfg_info order by form_id desc"
        cur.execute(sql)
        rows = cur.fetchall()
        form_list=[]
        for item in rows:
            form_list.append({"label":str(item[0])+'  '+str(item[1])+'  修改时间：'+str(item[2]), "value":str(item[0])} )

        cur.close()  # 关闭游标
    except Exception as ex:
        log.error("表单配置信息查询失败!" + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        public.respcode, public.respmsg = "100010", "表单配置信息查询失败!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    public.respcode, public.respmsg = "000000", "表单配置信息查询成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": {
            "form_list": form_list,
        }
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo


# 表单配置, 导出配置
def form_cfg_export(request):
    log = public.logger
    body = public.req_body
    form_id = body.get('form_id', '')
    try:
        cur = connection.cursor()  # 创建游标
        sql = "select form_id, form_name, update_date from sys_form_cfg_info order by form_id desc"
        cur.execute(sql)
        rows = cur.fetchall()
        form_list=[]
        for item in rows:
            form_list.append({"label": str(item[0])+'  '+str(item[1])+'  修改时间：'+str(item[2]), "value": str(item[0])} )

        cur.close()  # 关闭游标

        output = BytesIO()
        output.seek(0)
        # wb = openpyxl.load_workbook(BytesIO(output.read()))
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        # 写入表头
        for j, head_label in enumerate(auth_table_head_names):
            ws.cell(1, j + 1, head_label)
        # 写入数据
        for i, item in enumerate(table_data):
            for j, (k, v) in enumerate(item.items()):
                ws.cell(i + 2, j + 1, v)
        wb.save(output)
        base64_data = base64.b64encode(output.getvalue())
        file_base64 = base64_data.decode()
        public.respcode, public.respmsg = "125800", "导出成功!"
        date_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        json_data = {
            "HEAD": public.resphead_setvalue(),
            "BODY": {
                "filename": f"{form_name}_{date_str}.xlsx",
                "filetype": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "filedata": file_base64,
            }
        }
        s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
        public.respinfo = HttpResponse(s)
        return public.respinfo

    except Exception as ex:
        log.error("表单配置信息查询失败!" + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        public.respcode, public.respmsg = "100010", "表单配置信息查询失败!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

def GetRealSQL(sql, form_var, extra_data):
    log = public.logger
    pattern = re.compile("\$\[(.*?)\]")
    sqlvar = pattern.findall(sql)
    for sqlitm in sqlvar:
        old = "$[" + sqlitm + "]"
        if sqlitm in form_var.keys():
            new = "'" + str(form_var.get(sqlitm)) + "'"
        else:
            new = "''"
        sql = sql.replace(old, new)
    log.info('real sql=' + str(sql), extra={'ptlsh': public.req_seq})
    sql = public.SqlKeywordConver(sql, None, extra_data)
    log.info('finally sql=' + str(sql), extra={'ptlsh': public.req_seq})
    return sql

# 表单中表格数据初始化,增加初始化表格数据的操作
def form_table_show( request ):
    log = public.logger
    body = public.req_body
    form_id = body.get('form_id')
    table_id = body.get('table_id')
    pageSize = body.get('pageSize', '10')
    currentPage = body.get('currentPage', 1)
    search = body.get('search')
    form_var = body.get('form_var')
    # 20221203新增
    extra_data = body.get('extra')
    log.info(f'extra_data={extra_data}')
    if not form_id:
        public.respcode, public.respmsg = "100210", "表单ID不可为空!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo
    if not table_id:
        public.respcode, public.respmsg = "100211", "表格ID不可为空!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo
    try:
        cur = connection.cursor()  # 创建游标
        sql = "select form_cfg from sys_form_cfg_info where form_id = %s"
        log.info(sql % form_id, extra={'ptlsh': public.req_seq})
        cur.execute(sql, form_id)
        row = cur.fetchone()
        if not row:
            cur.close()  # 关闭游标
            public.respcode, public.respmsg = "100212", "表单配置不存在!"
            public.respinfo = HttpResponse(public.setrespinfo())
            return public.respinfo
        Layout_list = json.loads(row[0])

        # 递归获取表单中指定的表格组件
        def GetCompFormLayout(Layout_list, table_id):
            comp_cfg = {}
            for comp_item in Layout_list:
                if comp_item.get('children'):
                    comp_cfg = GetCompFormLayout(comp_item['children'], table_id)
                    if not comp_cfg:
                        continue
                    else:
                        return comp_cfg
                elif comp_item['type'] == 'null':
                    continue
                elif comp_item['id']==table_id:
                    comp_cfg = comp_item
                    return comp_cfg
            return comp_cfg #没找到

        comp_cfg = GetCompFormLayout(Layout_list, table_id)
        log.info('comp_cfg='+str(comp_cfg), extra={'ptlsh': public.req_seq})
        if not comp_cfg:
            cur.close()  # 关闭游标
            public.respcode, public.respmsg = "100212", "表格配置信息不存在!"
            public.respinfo = HttpResponse(public.setrespinfo())
            return public.respinfo

        variable_name = comp_cfg['attrs']['variable']  # 表格变量名
        sqlselect = comp_cfg['attrs']['data_sql']
        sqlorder = comp_cfg['attrs']['orderby_manage']
        sqlwhere= comp_cfg['attrs']['where_manage']
        if not sqlwhere:
            sqlwhere=" where 1=1 "
        if search:
            for searchitem in search:
                # if searchitem.get('a') and searchitem.get('to') and searchitem.get('b'):
                #     sqlwhere = sqlwhere + " and %s %s '%s' " % (searchitem['a'], searchitem['to'], searchitem['b'])
                if searchitem.get('a') and searchitem.get('to') and searchitem.get('b'):
                    # 检查日期字符串格式是否匹配
                    if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$', searchitem['b']):
                        date_object = datetime.datetime.strptime(searchitem['b'], '%Y-%m-%d %H:%M:%S')
                        searchitem['b'] = date_object.date().strftime('%Y-%m-%d')
                    sqlwhere=sqlwhere+" and %s %s '%s' " % (searchitem['a'], searchitem['to'],searchitem['b'])

        position_start = (currentPage-1) * pageSize
        position_end = currentPage * pageSize
        sqllimit = " limit %s, %s" % (position_start, pageSize)

        #获取记录总数
        sql = "%s %s" % (sqlselect, sqlwhere)
        sql = GetRealSQL(sql, form_var, extra_data)

        # 20230130更新 将sql中的select关键字和from关键字转为小写，其余不变
        sql = re.sub(r"^select\s",'select ', sql, flags=re.IGNORECASE)
        sql = re.sub(r"\sselect\s",' select ', sql, flags=re.IGNORECASE)
        sql = re.sub(r"\sfrom\s",' from ', sql, flags=re.IGNORECASE)
        # sql = sql.lower()

        fieldlist = sql.split('select')[1].split(' from ')[0]
        sql = sql.replace(fieldlist, ' count(1) ', 1)
        log.info('通用表格记录数查询sql:' + sql, extra={'ptlsh': public.req_seq})
        cur.execute(sql)
        row = cur.fetchone()
        body['table_total'] = row[0]

        # 获取记录明细:
        sql = "%s %s %s %s" % (sqlselect, sqlwhere, sqlorder, sqllimit)
        selsql = GetRealSQL(sql, form_var, extra_data)
        log.info('通用表格查询sql:' + selsql, extra={'ptlsh': public.req_seq})
        cur.execute(selsql)
        rows = cur.fetchall()
        # 获取字段名
        field_names = [i[0] for i in cur.description]
        log.info('字段名:' + str(field_names), extra={'ptlsh': public.req_seq})

        table_data = []
        for item in rows:
            data_item = {}
            i = 0
            for fielditem in field_names:
                # log.info('fielditem='+str(fielditem), extra={'ptlsh': public.req_seq})
                try:
                    data_item[fielditem] = item[i]
                    i = i+1
                except Exception as ex:
                    log.error("获取记录明细查询失败!" + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
                    break   # HEAD字段数量和SQL查询的字段不一致时，后边的不管了。

            table_data.append(data_item)
        body['table_data'] = table_data

        dict_data = {}
        # 20201125注释 不判断是否有数据
        # if len(body['table_data']) > 0: #有数据，先把表头字段对应的数据字典查出来
        for fielditem in comp_cfg['attrs']['head']:
            # log.info('fielditem='+str(fielditem), extra={'ptlsh': public.req_seq})
            morecfg = fielditem.get('moreConfig')
            if morecfg and morecfg.get('data_dict'):  # 按钮不赋值
                try:
                    data_item = []
                    dictsql = morecfg.get('data_dict')
                    # log.info(f'dictsql={dictsql}')
                    dictsql = GetRealSQL(dictsql, form_var, extra_data)
                    # log.info(f'dictsql={dictsql}')
                    cur.execute(dictsql)
                    dictrows = cur.fetchall()
                    for dictitem in dictrows:
                        data_item.append({"key": dictitem[0], "value": dictitem[1]})

                    dict_data[fielditem['name']] = data_item
                except Exception as ex:
                    log.error("sql:%s, error:%s" % (dictsql, str(ex)), exc_info=True, extra={'ptlsh': public.req_seq})
                    pass
        body['dict_data'] = dict_data

        cur.close()  # 关闭游标

    except Exception as ex:
        log.error("表单配置信息查询失败!" + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        cur.close()  # 关闭游标
        public.respcode, public.respmsg = "100010", "表单配置信息查询失败!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    public.respcode, public.respmsg = "000000", "表单配置信息查询成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": body
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo


#表单中配置时, 返回公共数据字典列表
def form_table_dict_list( request ):
    log = public.logger
    body = public.req_body

    try:
        cur = connection.cursor()  # 创建游标
        sql = "select distinct DICT_NAME, DICT_SNOTE from sys_ywty_dict where dict_public_flag='Y'"
        cur.execute( sql )
        rows = cur.fetchall()

        dict_list=[]
        for item in rows:
            data_item={
                "key": item[0],
                "value": item[1],
            }
            dict_list.append(data_item)
        body['dict_list'] = dict_list
        cur.close()  # 关闭游标

    except Exception as ex:
        log.error("表单配置信息查询失败!" + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        cur.close()  # 关闭游标
        public.respcode, public.respmsg = "100010", "表单配置信息查询失败!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    public.respcode, public.respmsg = "000000", "表单配置信息查询成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": body
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo

#表单中配置时, 返回数据字典键值
def form_table_dict_info( request ):
    log = public.logger
    body = public.req_body
    dict_name = body.get('dict_name')
    if not dict_name:
        public.respcode, public.respmsg = "100281", "字典名称不可为空!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    try:
        cur = connection.cursor()  # 创建游标
        sql = "select dict_code,dict_target from sys_ywty_dict where dict_name=%s"
        cur.execute( sql, dict_name )
        rows = cur.fetchall()

        dict_info = []
        for item in rows:
            data_item = {
                "key": item[0],
                "value": item[1],
            }
            dict_info.append(data_item)
        body['dict_info'] = dict_info
        cur.close()  # 关闭游标

    except Exception as ex:
        log.error("表单配置信息查询失败!" + str(ex), exc_info=True, extra={'ptlsh': public.req_seq})
        public.exc_type, public.exc_value, public.exc_traceback = sys.exc_info()
        cur.close()  # 关闭游标
        public.respcode, public.respmsg = "100010", "表单配置信息查询失败!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    public.respcode, public.respmsg = "000000", "表单配置信息查询成功!"
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": body
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo


# 导出数据
def export_data(request):
    log = public.logger
    body = public.req_body
    form_id = body.get('form_id')
    table_id = body.get('table_id')
    search = body.get('search')
    form_var = body.get('form_var', {})
    # 20221203新增
    extra_data = body.get('extra')
    log.info(f'extra_data={extra_data}')

    log.info('我是通用导出excel函数', extra={'ptlsh': public.req_seq})
    if not form_id:
        public.respcode, public.respmsg = "100210", "表单ID不可为空!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo
    if not table_id:
        public.respcode, public.respmsg = "100211", "表格ID不可为空!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    cur = connection.cursor()  # 创建游标
    sql = "select form_cfg, form_name from sys_form_cfg_info where form_id = %s"
    # log.info(sql % form_id, extra={'ptlsh': public.req_seq})
    cur.execute(sql, form_id)
    row = cur.fetchone()
    if not row:
        cur.close()  # 关闭游标
        public.respcode, public.respmsg = "100212", "表单配置不存在!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo
    Layout_list = json.loads(row[0])
    form_name = row[1]

    # 获取表格字段可显示的权限
    auth_table_head_ids = []
    auth_table_head_names = []
    sql = "select distinct a.field_id, a.field_name " \
          "from sys_role_purv_table_field a, sys_user_role b " \
          "where a.ROLE_ID=b.ROLE_ID and a.field_type='table_head' and a.show_able='Y' " \
          "and a.FORM_ID='%s' and b.USER_ID='%s' " % (form_id, public.user_id)
    # log.info("sql1:" + sql)
    cur.execute(sql)
    rows = cur.fetchall()
    if not rows: # 没配置，为了不影响导出，默认使用全部字段权限。
        sql = "select distinct field_id, field_name from sys_form_table_field " \
              "where field_type='table_head' and show_able='Y' and form_id='%s'" % form_id
        # log.info("sql2:" + sql)
        cur.execute(sql)
        rows = cur.fetchall()
    for item in rows:
        auth_table_head_ids.append(item[0])
        auth_table_head_names.append(item[1])  # 2025.8.5修改 by mwz
    log.info(f'auth_table_head_ids:{str(auth_table_head_ids)}', extra={'ptlsh': public.req_seq})
    log.info(f'auth_table_head_names:{auth_table_head_names}')

    # 递归获取表单中指定的表格组件
    def GetCompFormLayout(Layout_list, table_id):
        comp_cfg = {}
        for comp_item in Layout_list:
            if comp_item.get('children'):
                comp_cfg = GetCompFormLayout(comp_item['children'], table_id)
                if not comp_cfg:
                    continue
                else:
                    return comp_cfg
            elif comp_item['type'] == 'null':
                continue
            elif comp_item['id'] == table_id:
                comp_cfg = comp_item
                return comp_cfg
        return comp_cfg  # 没找到

    comp_cfg = GetCompFormLayout(Layout_list, table_id)
    log.info('comp_cfg=' + str(comp_cfg), extra={'ptlsh': public.req_seq})
    if not comp_cfg:
        cur.close()  # 关闭游标
        public.respcode, public.respmsg = "100212", "表格配置信息不存在!"
        public.respinfo = HttpResponse(public.setrespinfo())
        return public.respinfo

    sqlselect = comp_cfg['attrs']['data_sql']
    sqlorder = comp_cfg['attrs']['orderby_manage']
    sqlwhere = comp_cfg['attrs']['where_manage']

    if not sqlwhere:
        sqlwhere = " where 1=1 "
    if search:
        for searchitem in search:
            if searchitem.get('a') and searchitem.get('to') and searchitem.get('b'):
                sqlwhere = sqlwhere + " and %s %s '%s' " % (searchitem['a'], searchitem['to'], searchitem['b'])

    # 获取记录总数
    sql = "%s %s" % (sqlselect, sqlwhere)
    sql = GetRealSQL(sql, form_var, extra_data)

    # 20230130更新 将sql中的select关键字和from关键字转为小写，其余不变
    sql = re.sub(r"^select\s", 'select ', sql, flags=re.IGNORECASE)
    sql = re.sub(r"\sselect\s", ' select ', sql, flags=re.IGNORECASE)
    sql = re.sub(r"\sfrom\s", ' from ', sql, flags=re.IGNORECASE)
    # sql = sql.lower()

    # fieldlist = sql.split('select')[1].split(' from ')[0]
    # sql = sql.replace(fieldlist, ' count(1) ', 1)
    # log.info('通用表格记录数查询sql:' + sql, extra={'ptlsh': public.req_seq})
    # cur.execute(sql)
    # row = cur.fetchone()
    # body['table_total'] = row[0]

    # 获取记录明细:
    sql = "%s %s %s" % (sqlselect, sqlwhere, sqlorder)
    selsql = GetRealSQL(sql, form_var, extra_data)
    log.info('通用表格查询sql:' + selsql, extra={'ptlsh': public.req_seq})
    cur.execute(selsql)
    rows = cur.fetchall()

    # 获取字段名
    field_names = [i[0] for i in cur.description]
    log.info(f'查询语句中的字段名:{str(field_names)}', extra={'ptlsh': public.req_seq})

    table_data = []
    for item in rows:
        data_item = {}
        i = 0
        for fielditem in field_names:
            # log.info('fielditem='+str(fielditem), extra={'ptlsh': public.req_seq})
            try:
                i = i + 1
                if fielditem not in auth_table_head_ids:
                    continue
                data_item[fielditem] = item[i-1]
                # kv_tran = fielditem.get('moreConfig', {}).get('kv_tran')
                # if kv_tran and len(kv_tran) > 0:
                #     for tran_item in kv_tran:
                #         if str(item[i-1]) == tran_item['key']:
                #             data_item[fielditem] = tran_item['value']
                #             break
            except Exception as ex:
                log.error(str(ex), exc_info=True)
                break  # HEAD字段数量和SQL查询的字段不一致时，后边的不管了。

        table_data.append(data_item)
    log.info(f'table_data:{str(table_data)}', extra={'ptlsh': public.req_seq})

    # 20220906更新 后端数据字典转换 ---beigin
    dict_data = {}
    for fielditem in comp_cfg['attrs']['head']:
        # log.info('fielditem='+str(fielditem), extra={'ptlsh': public.req_seq})
        morecfg = fielditem.get('moreConfig')
        if morecfg and morecfg.get('data_dict'):  # 按钮不赋值
            try:
                data_item = []
                dictsql = morecfg.get('data_dict')
                cur.execute(dictsql)
                dictrows = cur.fetchall()
                for dictitem in dictrows:
                    data_item.append({"key": dictitem[0], "value": dictitem[1]})

                dict_data[fielditem['name']] = data_item
            except Exception as ex:
                log.info(str(ex))
                pass
    # body['dict_data'] = dict_data
    new_table_data = []
    # log.info(f'table_data={table_data}')
    # log.info(f'dict_data={dict_data}')
    for i, item in enumerate(table_data):
        for k, v in item.items():
            k_dict_data = dict_data.get(k)
            if not k_dict_data:
                continue
            for sub_item in k_dict_data:
                sub_key = sub_item['key']
                sub_value = sub_item['value']
                if str(v) == str(sub_key):
                    table_data[i][k] = sub_value

    # 20220906更新 后端数据字典转换 ---end

    output = BytesIO()
    output.seek(0)
    # wb = openpyxl.load_workbook(BytesIO(output.read()))
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    # 写入表头
    for j, head_label in enumerate(auth_table_head_names):
        ws.cell(1, j + 1, head_label)
    # 写入数据
    for i, item in enumerate(table_data):
        for j, (k, v) in enumerate(item.items()):
            ws.cell(i + 2, j + 1, v)
    wb.save(output)
    base64_data = base64.b64encode(output.getvalue())
    file_base64 = base64_data.decode()
    public.respcode, public.respmsg = "125800", "导出成功!"
    date_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    json_data = {
        "HEAD": public.resphead_setvalue(),
        "BODY": {
            "filename": f"{form_name}_{date_str}.xlsx",
            "filetype": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "filedata": file_base64,
        }
    }
    s = json.dumps(json_data, cls=public.JsonCustomEncoder, ensure_ascii=False)
    public.respinfo = HttpResponse(s)
    return public.respinfo