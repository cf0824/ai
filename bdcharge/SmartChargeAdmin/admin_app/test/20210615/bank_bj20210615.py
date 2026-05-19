# 中行报价单 excel 自动排日期
# import xlrd
# import xlwt

import openpyxl

book = openpyxl.load_workbook('授信合同管理系统报价单20210615.xlsx')
sheet = book.worksheets[0]
begin = -1
end = -1
for i in range(39):
    if i<4:
        continue
    if i>38:
        break

    p_num = float(sheet.cell(i,14).value)
    p_day = float(sheet.cell(i,15).value)
    print(p_num,p_day)
    day = int(p_day/p_num)
    begin = end+1
    end = begin+day-1
    begin_str = 'T+%s'%begin
    end_str = 'T+%s'%end
    begin_str = begin_str.replace('+0','')
    end_str = end_str.replace('+0','')
    sheet.cell(i,12,begin_str)
    sheet.cell(i,13,end_str)
    print(begin_str,end_str)

book.save('res.xlsx')
